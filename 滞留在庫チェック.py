import pandas as pd
import os
import sys
import shutil
import subprocess
from datetime import datetime, timedelta
import getpass
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import threading

# Application version
APP_VERSION = "1.0.3"
APP_NAME = "滞留在庫チェック"

# -----------------------------------------------------------------------------
# Auto-update functions
# -----------------------------------------------------------------------------

def get_update_folder_path():
    """Return the first accessible update folder path, or None."""
    username = getpass.getuser()
    candidate_paths = [
        f"C:\\Users\\{username}\\OneDrive - 東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\業務用pythonアプリ最新版\\滞留在庫チェック\\update",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\業務用pythonアプリ最新版\\滞留在庫チェック\\update",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部 - CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\業務用pythonアプリ最新版\\滞留在庫チェック\\update",
    ]
    for path in candidate_paths:
        if os.path.exists(path):
            print(f"DEBUG: アップデート用フォルダが見つかりました: {path}")
            return path
    print("DEBUG: アップデート用フォルダが見つかりませんでした。")
    return None


def compare_versions(version1, version2):
    """Return positive if version1 > version2, 0 if equal, negative if less."""
    def parse_version(v):
        return [int(x) for x in v.split('.')]

    v1_parts = parse_version(version1)
    v2_parts = parse_version(version2)
    for i in range(max(len(v1_parts), len(v2_parts))):
        v1_val = v1_parts[i] if i < len(v1_parts) else 0
        v2_val = v2_parts[i] if i < len(v2_parts) else 0
        if v1_val > v2_val:
            return 1
        elif v1_val < v2_val:
            return -1
    return 0


def get_latest_version():
    """Read version.txt from the update folder and return the version string."""
    update_folder = get_update_folder_path()
    if not update_folder:
        return None
    version_file = os.path.join(update_folder, "version.txt")
    if not os.path.exists(version_file):
        print(f"DEBUG: version.txtが見つかりません: {version_file}")
        return None
    try:
        with open(version_file, 'r', encoding='utf-8') as f:
            version = f.read().strip()
            print(f"DEBUG: 最新バージョン: {version}")
            return version
    except Exception as e:
        print(f"DEBUG: version.txt読み取りエラー: {e}")
        return None


def perform_update():
    """Copy the new script from the update folder and restart the app."""
    update_folder = get_update_folder_path()
    if not update_folder:
        messagebox.showerror("エラー", "アップデート用フォルダが見つかりません。")
        return False

    current_script = os.path.abspath(__file__)
    update_source = os.path.join(update_folder, os.path.basename(current_script))

    if not os.path.exists(update_source):
        messagebox.showerror("エラー", f"アップデートファイルが見つかりません:\n{update_source}")
        return False

    try:
        shutil.copy2(update_source, current_script)
        messagebox.showinfo("アップデート完了", "アップデートが完了しました。\nアプリケーションを再起動します。")
        subprocess.Popen([sys.executable, current_script, "--just-updated"])
        return True
    except Exception as e:
        messagebox.showerror("エラー", f"アップデート中にエラーが発生しました:\n{str(e)}")
        return False


def check_for_updates():
    """Check for a newer version at startup and prompt the user if found."""
    if "--just-updated" in sys.argv:
        print("DEBUG: アップデート直後のためチェックをスキップします。")
        return

    latest_version = get_latest_version()
    if latest_version is None:
        print("DEBUG: バージョン情報を取得できませんでした。アップデートチェックをスキップします。")
        return

    if compare_versions(latest_version, APP_VERSION) <= 0:
        print(f"DEBUG: 最新バージョンです。現在: {APP_VERSION}, 最新: {latest_version}")
        return

    print(f"DEBUG: 新しいバージョンが利用可能です。現在: {APP_VERSION}, 最新: {latest_version}")

    result = messagebox.askyesno(
        "アップデート確認",
        f"新しいバージョン {latest_version} が利用可能です。\n"
        f"現在のバージョン: {APP_VERSION}\n\n"
        "今すぐアップデートしますか？"
    )
    if result:
        if perform_update():
            sys.exit(0)


# -----------------------------------------------------------------------------

def get_file_path(file_patterns):
    """
    複数のファイルパスから存在するものを取得
    """
    current_user = getpass.getuser()
    
    for pattern in file_patterns:
        # [username]を現在のユーザー名に置換
        file_path = pattern.replace('[username]', current_user)
        if os.path.exists(file_path):
            return file_path
    
    raise FileNotFoundError(f"指定されたファイルが見つかりません")

def read_csv_with_encoding(file_path):
    """
    適切なエンコーディングでCSVファイルを読み込む
    """
    encodings = ['shift-jis', 'utf-8-sig', 'utf-8', 'cp932']
    
    for encoding in encodings:
        try:
            df = pd.read_csv(file_path, encoding=encoding)
            print(f"エンコーディング '{encoding}' で読み込み成功")
            return df
        except UnicodeDecodeError:
            continue
    
    raise UnicodeDecodeError(f"ファイルの読み込みに失敗しました: {file_path}")

def create_print_sheet(workbook, employee_name, sorted_df, month_input):
    """
    VBAのPrintInventoryReport相当の印刷用シートをopenpyxlで生成する
    """
    # Sheet name: max 31 chars
    print_sheet_name = f"印刷用_{employee_name}"[:31]
    ws = workbook.create_sheet(title=print_sheet_name)

    # 12 columns to extract (matching VBA column order)
    print_columns = [
        '入荷日', '経過日数', '受注番号', '受注件名', '受注日',
        '得意先略名', '商品コード', '商品略名', '入荷済数量',
        '発注金額', '対応', '出荷時期'
    ]
    print_df = sorted_df[print_columns].copy()

    # Calculate total order amount
    total_amount = sorted_df['発注金額'].sum()
    report_title = f"長期（3ヶ月以上）滞留在庫管理表_{month_input}"

    # Header section (rows 1-4)
    ws.cell(1, 1).value = report_title
    ws.cell(1, 1).font = Font(name="メイリオ", bold=True, size=14)

    ws.cell(2, 1).value = f"担当者名：{employee_name}"
    ws.cell(2, 1).font = Font(name="メイリオ", bold=True, size=10)

    ws.cell(3, 1).value = "合計発注金額："
    ws.cell(3, 1).font = Font(name="メイリオ", bold=True, size=10)
    ws.cell(3, 2).value = total_amount
    ws.cell(3, 2).number_format = '#,##0'
    ws.cell(3, 2).font = Font(name="メイリオ", bold=True, size=10)

    DATA_START_ROW = 5

    # Column header row (row 5)
    for col_idx, col_name in enumerate(print_columns, start=1):
        cell = ws.cell(DATA_START_ROW, col_idx)
        cell.value = col_name
        cell.font = Font(name="メイリオ", bold=True, size=10)

    # Data rows (row 6+): alternate blue/white by 受注番号
    light_blue_fill = PatternFill(start_color='E6F2FF', end_color='E6F2FF', fill_type='solid')
    previous_order_number = None
    use_color = False

    for row_idx, (_, row_data) in enumerate(print_df.iterrows()):
        excel_row = DATA_START_ROW + 1 + row_idx
        current_order_number = row_data['受注番号']

        if current_order_number != previous_order_number:
            use_color = not use_color
            previous_order_number = current_order_number

        for col_idx, col_name in enumerate(print_columns, start=1):
            cell = ws.cell(excel_row, col_idx)
            cell.value = row_data[col_name]
            cell.font = Font(name="メイリオ", size=10)
            if use_color:
                cell.fill = light_blue_fill
            if col_name == '発注金額':
                cell.number_format = '#,##0'

    # Borders for the entire data range (header row + data rows)
    last_data_row = DATA_START_ROW + len(print_df)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=last_data_row,
                             min_col=1, max_col=12):
        for cell in row:
            cell.border = border

    # Column widths
    # Fixed widths: A/B/E/F = 100px → 13.57, D = 145px → 20.0, K/L = 185px → 25.71
    fixed_widths = {'A': 13.57, 'B': 13.57, 'D': 20.0, 'E': 13.57, 'F': 13.57, 'K': 25.71, 'L': 25.71}
    for col_idx, col_name in enumerate(print_columns, start=1):
        col_letter = get_column_letter(col_idx)
        if col_letter in fixed_widths:
            ws.column_dimensions[col_letter].width = fixed_widths[col_letter]
        else:
            header_len = len(col_name)
            if len(print_df) > 0:
                data_len = print_df[col_name].astype(str).str.len().max()
            else:
                data_len = 0
            ws.column_dimensions[col_letter].width = min(max(header_len, data_len) + 2, 50)

    # AutoFilter on the column header row
    ws.auto_filter.ref = f"A{DATA_START_ROW}:L{DATA_START_ROW}"

    # Page setup (landscape, B4, fit to 1 page wide)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 12  # B4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:L{last_data_row}"
    ws.print_title_rows = f"1:{DATA_START_ROW}"

    # Footer: page number
    ws.oddFooter.center.text = "&P/&N"

    # Margins (inches)
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1


def process_data(days_threshold, month_input, progress_callback=None):
    """
    データ処理のメイン関数
    """
    # ファイルパスの定義
    nyuka_patterns = [
        r"C:\Users\[username]\OneDrive - 東邦ヤンマーテック株式会社\CR推進本部フォルダ\06_社内管理資料\miraimiru移行関連\フォルダ共有テスト\入荷実績\【標準】_入荷.csv",
        r"C:\Users\[username]\東邦ヤンマーテック株式会社\CR推進本部フォルダ\06_社内管理資料\miraimiru移行関連\フォルダ共有テスト\入荷実績\【標準】_入荷.csv",
        r"C:\Users\[username]\東邦ヤンマーテック株式会社\CR推進本部 - CR推進本部フォルダ\06_社内管理資料\miraimiru移行関連\フォルダ共有テスト\入荷実績\【標準】_入荷.csv"
    ]
    
    shukka_patterns = [
        r"C:\Users\[username]\OneDrive - 東邦ヤンマーテック株式会社\CR推進本部フォルダ\06_社内管理資料\miraimiru移行関連\フォルダ共有テスト\出荷実績\【標準】_出荷.csv",
        r"C:\Users\[username]\東邦ヤンマーテック株式会社\CR推進本部フォルダ\06_社内管理資料\miraimiru移行関連\フォルダ共有テスト\出荷実績\【標準】_出荷.csv",
        r"C:\Users\[username]\東邦ヤンマーテック株式会社\CR推進本部 - CR推進本部フォルダ\06_社内管理資料\miraimiru移行関連\フォルダ共有テスト\出荷実績\【標準】_出荷.csv"
    ]
    
    juchuu_patterns = [
        r"C:\Users\[username]\OneDrive - 東邦ヤンマーテック株式会社\CR推進本部フォルダ\06_社内管理資料\miraimiru移行関連\フォルダ共有テスト\受注残\【標準】_受注.csv",
        r"C:\Users\[username]\東邦ヤンマーテック株式会社\CR推進本部フォルダ\06_社内管理資料\miraimiru移行関連\フォルダ共有テスト\受注残\【標準】_受注.csv",
        r"C:\Users\[username]\東邦ヤンマーテック株式会社\CR推進本部 - CR推進本部フォルダ\06_社内管理資料\miraimiru移行関連\フォルダ共有テスト\受注残\【標準】_受注.csv"
    ]
    
    try:
        # ファイルパスを取得
        if progress_callback:
            progress_callback("ファイルパスを取得中...")
        
        nyuka_path = get_file_path(nyuka_patterns)
        shukka_path = get_file_path(shukka_patterns)
        juchuu_path = get_file_path(juchuu_patterns)
        
        print(f"入荷ファイル: {nyuka_path}")
        print(f"出荷ファイル: {shukka_path}")
        print(f"受注ファイル: {juchuu_path}")
        
        # 1. 入荷データの読み込みと処理
        if progress_callback:
            progress_callback("1. 入荷データを処理中...")
        print("\n1. 入荷データを処理中...")
        nyuka_df = read_csv_with_encoding(nyuka_path)
        
        # 入荷日をdatetime型に変換
        nyuka_df['入荷日'] = pd.to_datetime(nyuka_df['入荷日'])
        
        # 指定された日数以上前の日付を計算
        today = datetime.now()
        threshold_date = today - timedelta(days=days_threshold)
        
        # 指定された日数以上前のレコードを抽出
        nyuka_filtered = nyuka_df[nyuka_df['入荷日'] <= threshold_date].copy()
        
        if nyuka_filtered.empty:
            message = f"{days_threshold}日以上前の入荷データが存在しません。"
            print(message)
            if progress_callback:
                progress_callback(message)
            return
        
        # 主キーを作成（明細_共通項目3と明細_共通項目2の連結）
        nyuka_filtered['主キー'] = nyuka_filtered['明細_共通項目3'].astype(str) + '_' + nyuka_filtered['明細_共通項目2'].astype(str)
        
        # 主キーごとに明細_入荷数量を集計
        nyuka_qty_sum = nyuka_filtered.groupby('主キー')['明細_入荷数量'].sum().to_dict()
        
        # 必要な列を抽出
        nyuka_data = nyuka_filtered[['主キー', '入荷日', '明細_ロット番号']].copy()
        
        print(f"{days_threshold}日以上前の入荷データ: {len(nyuka_data)}件")
        
        # 2. 出荷データの読み込みとチェック
        if progress_callback:
            progress_callback("2. 出荷データをチェック中...")
        print("\n2. 出荷データをチェック中...")
        shukka_df = read_csv_with_encoding(shukka_path)
        
        # 出荷データのロット番号を取得
        shukka_lot_numbers = set(shukka_df['明細_ロット番号'].dropna().astype(str))
        
        # 入荷データのロット番号が出荷データに含まれているかチェック
        nyuka_data['出荷済み'] = nyuka_data['明細_ロット番号'].astype(str).isin(shukka_lot_numbers)
        
        # 出荷されていないデータを抽出
        not_shipped = nyuka_data[~nyuka_data['出荷済み']].copy()
        
        print(f"出荷されていないデータ: {len(not_shipped)}件")
        
        if not_shipped.empty:
            message = "全てのデータが出荷済みです。"
            print(message)
            if progress_callback:
                progress_callback(message)
            return
        
        # 3. 受注データの読み込みと結合
        if progress_callback:
            progress_callback("3. 受注データを処理中...")
        print("\n3. 受注データを処理中...")
        juchuu_df = read_csv_with_encoding(juchuu_path)
        
        # 受注データの主キーを作成
        juchuu_df['主キー'] = juchuu_df['明細_共通項目3'].astype(str) + '_' + juchuu_df['明細_共通項目2'].astype(str)

        # 受注データから見積番号・行番号が空欄(NaN)のデータを除外
        juchuu_filtered = juchuu_df[
            juchuu_df['明細_共通項目3'].notna() &
            juchuu_df['明細_共通項目2'].notna() &
            (juchuu_df['明細_共通項目3'] != 'nan') &
            (juchuu_df['明細_共通項目2'] != 'nan')
        ].copy()

        print(f"受注データフィルタリング前: {len(juchuu_df)}件")
        print(f"受注データフィルタリング後: {len(juchuu_filtered)}件")

        # 未出荷のデータと受注データを結合
        merged_data = not_shipped.merge(juchuu_filtered, on='主キー', how='left')

        # デバッグ: マッチしなかったデータを確認
        unmatched = merged_data[merged_data['明細_受注残金額'].isna()]
        if not unmatched.empty:
            print(f"\n【デバッグ】受注データとマッチしなかったデータ: {len(unmatched)}件")
            print("サンプル(最初の5件):")
            print(unmatched[['主キー', '入荷日', '明細_ロット番号']].head())

        # 明細_受注残金額が0でないデータを抽出（受注データとマッチしなかったデータも除外）
        result_data = merged_data[
            (merged_data['明細_受注残金額'].notna()) &  # 受注データとマッチしたもの
            (merged_data['明細_受注残金額'] != 0)        # かつ受注残金額が0でないもの
        ].copy()

        print(f"受注データとマッチして受注残金額が0でないデータ: {len(result_data)}件")
        
        if result_data.empty:
            message = "受注残金額が0でないデータが存在しません。"
            print(message)
            if progress_callback:
                progress_callback(message)
            return
        
        # 明細_商品コードによるフィルタリング
        # 空欄、888888-88888、777777-77777のレコードを除外
        filtered_data = result_data[
            ~result_data['明細_商品コード'].isnull() &  # 空欄を除外
            (result_data['明細_商品コード'] != '888888-88888') &  # 888888-88888を除外
            (result_data['明細_商品コード'] != '777777-77777')    # 777777-77777を除外
        ].copy()
        
        print(f"商品コードフィルタリング前: {len(result_data)}件")
        print(f"商品コードフィルタリング後: {len(filtered_data)}件")
        
        if filtered_data.empty:
            message = "フィルタリング後のデータが存在しません。"
            print(message)
            if progress_callback:
                progress_callback(message)
            return
        
        # 入荷済数量を主キーから取得
        filtered_data['入荷済数量'] = filtered_data['主キー'].map(nyuka_qty_sum).fillna(0)
        
        # 発注金額を計算（明細_共通項目1 × 入荷済数量）
        filtered_data['発注金額'] = filtered_data['明細_共通項目1'] * filtered_data['入荷済数量']
        
        # 経過日数を計算（入荷日から今日までの日数）
        filtered_data['経過日数'] = (today - filtered_data['入荷日']).dt.days
        
        # 新しい列を追加（対応、出荷時期）
        filtered_data['対応'] = ''  # 空の列を追加
        filtered_data['出荷時期'] = ''  # 空の列を追加
        
        # 新しい列順序を定義（対応、出荷時期を追加）
        output_columns = [
            '入荷日',
            '経過日数',
            '受注番号',
            '明細_共通項目3',  # ← 新規追加（見積番号）
            '受注件名',
            '受注日',
            '部門名',
            '社員名',
            '得意先',
            '得意先略名',
            '受渡場所名',
            '明細_商品コード',
            '明細_商品略名',
            '明細_受注数量',
            '入荷済数量',
            '明細_受注単価',
            '明細_明細金額',
            '明細_共通項目1',
            '発注金額',
            '明細_共通項目2',
            '対応',
            '出荷時期'
        ]
        
        # 出力用データフレームを作成（列順序を指定）
        output_df = filtered_data[output_columns].copy()

        # デバッグ: 見積番号が空欄のデータを確認
        empty_mitumori = output_df[output_df['明細_共通項目3'].isna() | (output_df['明細_共通項目3'] == 'nan')]
        if not empty_mitumori.empty:
            print(f"\n【デバッグ】見積番号が空欄のデータ: {len(empty_mitumori)}件")
            print("サンプル(最初の3件):")
            print(empty_mitumori[['入荷日', '明細_共通項目3', '明細_共通項目2', '明細_商品コード', '明細_商品略名']].head(3))

        # 列名を変更（明細_を削除 + 共通項目の変更）
        column_mapping = {
            '明細_商品コード': '商品コード',
            '明細_商品略名': '商品略名',
            '明細_受注数量': '受注数量',
            '明細_受注単価': '受注単価',
            '明細_明細金額': '受注金額',
            '明細_共通項目1': '発注単価',
            '明細_共通項目2': '行番号',
            '明細_共通項目3': '見積番号'  # ← 新規追加
        }
        output_df = output_df.rename(columns=column_mapping)
        
        # 入荷日を年月日のみの形式に変換
        output_df['入荷日'] = output_df['入荷日'].dt.date
        
        # 出力ファイル名を生成
        output_filename = f"未出荷在庫_受注残_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # ドキュメントフォルダのパスを取得
        documents_path = os.path.join(os.path.expanduser('~'), 'Documents')
        output_path = os.path.join(documents_path, output_filename)
        
        # Excelファイルに出力（社員名ごとにシートを分ける）
        if progress_callback:
            progress_callback("Excelファイルを作成中...")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 社員名ごとの発注金額を集計
            summary_df = output_df.groupby('社員名')['発注金額'].sum().reset_index()
            summary_df = summary_df.sort_values('発注金額', ascending=False)
            
            # 全社計の行を追加
            total_row = pd.DataFrame([{
                '社員名': '全社計',
                '発注金額': summary_df['発注金額'].sum()
            }])
            summary_df = pd.concat([summary_df, total_row], ignore_index=True)
            
            # 集計シートを作成
            summary_df.to_excel(writer, sheet_name='集計', index=False)
            
            # 集計シートの書式設定
            summary_worksheet = writer.sheets['集計']
            summary_worksheet.freeze_panes = 'A2'
            
            # 列幅の自動調整と数値フォーマット設定
            for idx, col in enumerate(summary_df.columns):
                column_letter = get_column_letter(idx + 1)
                
                # 列の最大文字数を取得
                header_length = len(col)
                
                if col == '発注金額':
                    # 数値の場合はカンマを含めた文字数を計算
                    data_length = summary_df[col].apply(lambda x: len(f"{x:,.0f}") if pd.notna(x) else 0).max()
                else:
                    data_length = summary_df[col].astype(str).str.len().max()
                
                max_length = max(header_length, data_length) + 2
                
                # 列幅を設定
                summary_worksheet.column_dimensions[column_letter].width = min(max_length, 50)
                
                # カンマ区切りの数値フォーマットを設定
                if col == '発注金額':
                    for row in range(2, len(summary_df) + 2):
                        cell = summary_worksheet[f"{column_letter}{row}"]
                        cell.number_format = '#,##0'
            
            # 全社計の行を太字・上部ボーダーで装飾
            last_row = len(summary_df) + 1
            for col in range(1, len(summary_df.columns) + 1):
                cell = summary_worksheet.cell(row=last_row, column=col)
                cell.font = Font(bold=True)
                # 上部にボーダーを追加
                border = Border(top=Side(style='thin'))
                cell.border = border
            
            # 社員名でグループ化してシートを作成
            for employee_name, group_df in output_df.groupby('社員名'):
                # データをソート（得意先、受注番号、行番号の昇順）
                sorted_df = group_df.sort_values(['得意先', '受注番号', '行番号'], ascending=[True, True, True])
                
                # シート名を作成（ExcelのシートはA31文字まで）
                sheet_name = str(employee_name)[:31] if employee_name else "未登録"
                
                # データを出力
                sorted_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # ワークシートの書式設定
                worksheet = writer.sheets[sheet_name]
                
                # ヘッダー行を固定
                worksheet.freeze_panes = 'A2'
                
                # オートフィルタを設定
                last_column_letter = get_column_letter(len(sorted_df.columns))
                last_row = len(sorted_df) + 1
                worksheet.auto_filter.ref = f"A1:{last_column_letter}{last_row}"
                
                # カンマ区切りにする列の定義
                comma_columns = ['受注単価', '受注金額', '発注単価', '発注金額']
                
                # 受注番号ごとの行の色付け用の定義
                light_blue_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
                
                # 黄色の背景色を定義（「対応」「出荷時期」ヘッダー用）
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                
                # 受注番号でグループを判定するための準備
                previous_order_number = None
                use_color = False
                
                # 列幅の自動調整、数値フォーマット設定、行の色付け
                for idx, col in enumerate(sorted_df.columns):
                    column_letter = get_column_letter(idx + 1)
                    
                    # 列の最大文字数を取得（ヘッダーとデータの最大値）
                    header_length = len(col)
                    
                    # データの最大文字数を計算（カンマ区切りを考慮）
                    if col in comma_columns:
                        # 数値の場合はカンマを含めた文字数を計算
                        data_length = sorted_df[col].apply(lambda x: len(f"{x:,.0f}") if pd.notna(x) else 0).max()
                    else:
                        data_length = sorted_df[col].astype(str).str.len().max()
                    
                    max_length = max(header_length, data_length) + 2
                    
                    # 列幅を設定（最大100文字）
                    worksheet.column_dimensions[column_letter].width = min(max_length, 100)
                    
                    # ヘッダー行の「対応」「出荷時期」を黄色で色付け
                    if col in ['対応', '出荷時期']:
                        header_cell = worksheet[f"{column_letter}1"]
                        header_cell.fill = yellow_fill
                    
                    # カンマ区切りの数値フォーマットを設定
                    if col in comma_columns:
                        for row in range(2, len(sorted_df) + 2):  # データ行（ヘッダーを除く）
                            cell = worksheet[f"{column_letter}{row}"]
                            cell.number_format = '#,##0'
                
                # 受注番号ごとに行の色を設定
                for row_idx in range(len(sorted_df)):
                    excel_row = row_idx + 2  # Excelの行番号（ヘッダーを除く）
                    current_order_number = sorted_df.iloc[row_idx]['受注番号']
                    
                    # 受注番号が変わったら色を切り替え
                    if current_order_number != previous_order_number:
                        use_color = not use_color
                        previous_order_number = current_order_number
                    
                    # 色を適用
                    if use_color:
                        for col in range(1, len(sorted_df.columns) + 1):
                            cell = worksheet.cell(row=excel_row, column=col)
                            cell.fill = light_blue_fill

            # 印刷用シートを全担当者分作成
            if progress_callback:
                progress_callback("印刷用シートを作成中...")
            for employee_name, group_df in output_df.groupby('社員名'):
                sorted_df_for_print = group_df.sort_values(
                    ['得意先', '受注番号', '行番号'], ascending=[True, True, True]
                )
                create_print_sheet(writer.book, employee_name, sorted_df_for_print, month_input)

            # 担当者別の元シートを削除
            employee_names = output_df['社員名'].unique()
            for employee_name in employee_names:
                sheet_name = str(employee_name)[:31] if employee_name else "未登録"
                if sheet_name in writer.book.sheetnames:
                    del writer.book[sheet_name]

        success_message = f"処理完了！\n出力ファイル: {output_path}\n出力件数: {len(output_df)}件\n社員数: {output_df['社員名'].nunique()}名"
        print(f"\n{success_message}")
        if progress_callback:
            progress_callback(success_message)
        
        return output_path
        
    except FileNotFoundError as e:
        error_message = f"エラー: {e}"
        print(error_message)
        if progress_callback:
            progress_callback(error_message)
        raise e
    except Exception as e:
        error_message = f"予期しないエラーが発生しました: {e}"
        print(error_message)
        if progress_callback:
            progress_callback(error_message)
        import traceback
        traceback.print_exc()
        raise e

class MainGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("未出荷在庫・受注残 データ処理")
        self.root.geometry("400x320")

        # 日数入力の設定
        tk.Label(root, text="滞留日数を入力してください（日）：", font=("Arial", 12)).pack(pady=20)

        self.days_var = tk.StringVar(value="90")  # デフォルト値

        # 半角数字(0-9)のみ許可するバリデーション（全角数字はisdigit()でTrueになるため明示チェック）
        vcmd = (root.register(lambda val: all(c in '0123456789' for c in val)), '%P')

        days_entry_frame = tk.Frame(root)
        days_entry_frame.pack(pady=5)
        self.days_entry = tk.Entry(
            days_entry_frame,
            textvariable=self.days_var,
            validate='key',
            validatecommand=vcmd,
            font=("Arial", 12),
            width=8,
            justify='center'
        )
        self.days_entry.pack(side=tk.LEFT)
        tk.Label(days_entry_frame, text="日以上前", font=("Arial", 12)).pack(side=tk.LEFT, padx=5)
        tk.Label(root, text="※ 半角数字で入力してください", font=("Arial", 9), fg="gray").pack()
        
        # 月度入力欄
        tk.Label(root, text="月度を入力してください：", font=("Arial", 11)).pack(pady=(10, 0))
        self.month_var = tk.StringVar(value=f"{datetime.now().year}年{datetime.now().month}月度")
        tk.Entry(root, textvariable=self.month_var, font=("Arial", 11), width=20).pack(pady=5)

        # 実行ボタン
        self.execute_button = tk.Button(root, text="実行", command=self.execute_processing,
                                      font=("Arial", 12), bg="lightblue", width=15, height=2)
        self.execute_button.pack(pady=15)
        
        # 進捗表示
        self.progress_label = tk.Label(root, text="", font=("Arial", 10))
        self.progress_label.pack(pady=10)
    
    def execute_processing(self):
        # 月度のバリデーション
        month_input = self.month_var.get().strip()
        if not month_input:
            messagebox.showwarning("入力エラー", "月度を入力してください。")
            return

        # 滞留日数のバリデーション
        days_str = self.days_var.get().strip()
        if not days_str:
            messagebox.showwarning("入力エラー", "滞留日数を入力してください。")
            return
        days = int(days_str)
        if days < 1:
            messagebox.showwarning("入力エラー", "滞留日数は1以上の数値を入力してください。")
            return

        # ボタンを無効化
        self.execute_button.config(state="disabled")
        self.progress_label.config(text="処理を開始しています...")

        # 別スレッドで処理を実行
        thread = threading.Thread(target=self.run_processing, args=(days, month_input))
        thread.start()
    
    def run_processing(self, days, month_input):
        try:
            def update_progress(message):
                self.root.after(0, lambda: self.progress_label.config(text=message))

            output_path = process_data(days, month_input, progress_callback=update_progress)
            
            # 処理成功のメッセージ
            self.root.after(0, lambda: messagebox.showinfo("処理完了", 
                f"処理が完了しました。\nファイルが保存されました：\n{output_path}"))
        
        except Exception as e:
            # エラーメッセージ
            self.root.after(0, lambda: messagebox.showerror("エラー", 
                f"処理中にエラーが発生しました：\n{str(e)}"))
        
        finally:
            # ボタンを有効化
            self.root.after(0, lambda: self.execute_button.config(state="normal"))
            self.root.after(0, lambda: self.progress_label.config(text=""))

if __name__ == "__main__":
    # Auto-update check (requires a hidden Tk root for messagebox)
    temp_root = tk.Tk()
    temp_root.withdraw()
    check_for_updates()
    try:
        temp_root.destroy()
    except Exception:
        pass

    root = tk.Tk()
    app = MainGUI(root)
    root.mainloop()